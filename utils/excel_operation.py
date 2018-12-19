#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time


# 插入数据
# file 文件名
# sheet_name sheet 名
# header 表头
# 数据列表
# Attention:
# 一次性存大量数据时不要直接用该方法，因为会多次打开关闭文件，速度很慢
def write_data(file, sheet_name, headers, data):
    if not os.path.exists(file):
        Workbook().save(file)
    workbook = load_workbook(file)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name, len(workbook.sheetnames))
    sheet = workbook[sheet_name]
    print(sheet.max_row)
    if sheet.max_row == 1:
        sheet.append(headers)
    sheet.append(data)
    workbook.save(file)


if __name__ == '__main__':
    for i in range(10000):
        start = time.time()
        write_data("1.xlsx", "1", ["header0", "header1"], ["data1", "data2"])
        print(time.time() - start)
