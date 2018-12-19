#!/usr/bin/python
# -*- coding: utf-8 -*-
import os
import unittest
from openpyxl import Workbook

from openpyxl import load_workbook
from utils.db_helper import get_db_connect

settings = [{
    'table_name': 'hefeixin_2018',
    'product_name': ['和工作', '智联招聘'],
    'bussiness': ['start_delay', 'search_delay', 'ram', 'cpu', 'search_delay',
                  'detail_delay'],
    'network': [4, 6]  # 4为4G ,6为WiFi, 固网
}, ]


# 数据统计类
# 根据setting 的设置 循环查询数据， 放入Excel

class DataStatistics(unittest.TestCase):

    def test_data(self):
        conn = get_db_connect()
        cursor = conn.cursor()
        for setting in settings:
            table_name = setting['table_name']
            bussinesses = setting['bussiness']
            product_names = setting['product_name']
            networks = setting['network']
            file_name = './%s.xlsx' % table_name
            workbook, sheet = self.get_workbook_sheet(file_name, table_name)
            for product_name in product_names:
                for bussiness in bussinesses:
                    for network in networks:
                        sheet.append(self.get_data_values(conn, table_name, product_name, bussiness, network))
            workbook.save(file_name)
        cursor.close()

    # 获取对应产品 ，用例，网络的list
    # 头部加上产品 ，用例，网络，用于Excel识别
    def get_data_values(self, conn, table_name, product_name, bussiness, network):
        cursor = conn.cursor()
        sql = '''SELECT data_value FROM %s WHERE product_name = '%s' AND bussiness = '%s' AND network = "%s"''' % (
            table_name, product_name, bussiness, network)
        cursor.execute(sql)
        datas = cursor.fetchall()
        data_list = [product_name, bussiness, network]
        for data in datas:
            data_list.append(data[0])
        print(data_list)
        return data_list

    # 获取处理Excel的workbook与sheet对象
    def get_workbook_sheet(self, file, sheet_name):
        if not os.path.exists(file):
            Workbook().save(file)
        workbook = load_workbook(file)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name, len(workbook.sheetnames))
        sheet = workbook[sheet_name]
        return workbook, sheet


if __name__ == '__main__':
    unittest.main()
